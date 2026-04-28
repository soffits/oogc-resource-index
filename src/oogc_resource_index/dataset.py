from __future__ import annotations

import csv
from collections.abc import Iterable
from pathlib import Path

from openpyxl import Workbook, load_workbook

from oogc_resource_index.constants import DATASET_FIELDS, FID_FIELD
from oogc_resource_index.models import ResourceRecord


def sorted_records(records: Iterable[ResourceRecord]) -> list[ResourceRecord]:
    return sorted(records, key=lambda record: _fid_sort_key(record.fid))


def merge_records(
    existing: Iterable[ResourceRecord], updates: Iterable[ResourceRecord]
) -> list[ResourceRecord]:
    merged = {record.fid: record for record in existing if record.fid}
    for update in updates:
        if not update.fid:
            continue
        current = merged.get(update.fid)
        if current is None:
            merged[update.fid] = update
            continue
        row = current.as_row()
        for field, value in update.as_row().items():
            if value:
                row[field] = value
        merged[update.fid] = ResourceRecord.from_row(row)
    return sorted_records(merged.values())


def read_dataset(path: Path) -> list[ResourceRecord]:
    if not path.exists():
        return []
    if path.suffix.lower() == ".xlsx":
        return _read_xlsx(path)
    return _read_csv(path)


def write_dataset(records: Iterable[ResourceRecord], path: Path) -> Path:
    normalized = sorted_records(records)
    if path.suffix.lower() == ".xlsx":
        try:
            _write_xlsx(normalized, path)
            return path
        except Exception:
            fallback = path.with_suffix(".csv")
            _write_csv(normalized, fallback)
            return fallback
    _write_csv(normalized, path)
    return path


def write_csv_copy(records: Iterable[ResourceRecord], path: Path) -> Path:
    csv_path = path.with_suffix(".csv") if path.suffix.lower() == ".xlsx" else path
    _write_csv(sorted_records(records), csv_path)
    return csv_path


def _read_csv(path: Path) -> list[ResourceRecord]:
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        return [ResourceRecord.from_row(row) for row in csv.DictReader(file) if row.get(FID_FIELD)]


def _read_xlsx(path: Path) -> list[ResourceRecord]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    header = next(rows, None)
    if header is None:
        return []
    fields = [str(value or "").strip() for value in header]
    records: list[ResourceRecord] = []
    for values in rows:
        row = dict(zip(fields, values, strict=False))
        if row.get(FID_FIELD):
            records.append(ResourceRecord.from_row(row))
    workbook.close()
    return records


def _write_csv(records: list[ResourceRecord], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=DATASET_FIELDS)
        writer.writeheader()
        writer.writerows(record.as_row() for record in records)


def _write_xlsx(records: list[ResourceRecord], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "oogc_resources"
    worksheet.append(DATASET_FIELDS)
    for record in records:
        row = record.as_row()
        worksheet.append([row[field] for field in DATASET_FIELDS])
    workbook.save(path)


def _fid_sort_key(fid: str) -> tuple[int, str]:
    if fid.isdigit():
        return (0, f"{int(fid):020d}")
    return (1, fid)
